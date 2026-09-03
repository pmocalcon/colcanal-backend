# -*- coding: utf-8 -*-
"""Vuelca la hoja «Prestamos» de «01. Informe general de préstamos.xlsx» a JSON.

    python src/database/scripts/leer-hoja-prestamos.py <ruta-xlsx> <ruta-json-salida>

Solo lee el libro. El JSON que produce es lo que come
`actualizar-prestamos-desde-hoja.ts`, que es el que sí escribe en la base.

Se hace en dos pasos —y no todo en el script de TypeScript— para poder mirar el volcado
antes de que nada toque producción, y porque la hoja se revisa a fin de mes: el mismo
JSON sirve para el parte y para la carga.

La hoja tiene una retícula de 52 columnas, de septiembre de 2022 a diciembre de 2026 (K
a BJ), con una celda por mes. Acá se endereza a filas. Un 0 y una celda vacía son lo
mismo: ese mes no se descontó.
"""
import datetime
import io
import json
import sys

import openpyxl

MESES = []
_anio, _mes = 2022, 9
for _ in range(52):
    MESES.append((_anio, _mes))
    _mes += 1
    if _mes == 13:
        _mes, _anio = 1, _anio + 1


def txt(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() or None


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main(ruta_xlsx, ruta_json):
    ws = openpyxl.load_workbook(ruta_xlsx, data_only=True)["Prestamos"]

    filas = []
    for r in range(4, ws.max_row + 1):
        nombre = txt(ws.cell(r, 2).value)
        if not nombre:
            continue
        pagos = [
            {"anio": a, "mes": m, "valor": num(ws.cell(r, 11 + i).value)}
            for i, (a, m) in enumerate(MESES)
            if num(ws.cell(r, 11 + i).value)
        ]
        cuotas = num(ws.cell(r, 7).value)
        filas.append({
            "fila": r,
            "numero": int(ws.cell(r, 1).value) if isinstance(ws.cell(r, 1).value, (int, float)) else None,
            "nombre": nombre,
            "estado": txt(ws.cell(r, 3).value),
            "proyecto": txt(ws.cell(r, 4).value),
            "pagare": txt(ws.cell(r, 5).value),
            "mesInicio": txt(ws.cell(r, 6).value),
            "numeroCuotas": int(cuotas) if cuotas is not None else None,
            "fechaVencimiento": txt(ws.cell(r, 8).value),
            "valorPrestamo": num(ws.cell(r, 9).value),
            "valorCuota": num(ws.cell(r, 10).value),
            "valorCancelado": num(ws.cell(r, 63).value),
            "saldo": num(ws.cell(r, 64).value),
            "obs": " · ".join(x for x in [txt(ws.cell(r, 65).value), txt(ws.cell(r, 66).value)] if x) or None,
            "pagos": pagos,
            "sumaPagos": round(sum(p["valor"] for p in pagos), 2),
        })

    io.open(ruta_json, "w", encoding="utf-8").write(
        json.dumps({"filas": filas}, ensure_ascii=False, indent=1)
    )

    reales = [f for f in filas if not f["nombre"].upper().startswith(("TOTAL", "SALDO PENDIENTE"))]
    print(f"{len(reales)} prestamos · {sum(len(f['pagos']) for f in reales)} meses con valor")

    # Si la suma de los meses no da el VALOR CANCELADO, la hoja se contradice a sí misma
    # y hay que resolverlo con Contabilidad antes de cargar nada.
    descuadre = [
        f for f in reales
        if f["valorCancelado"] is not None and abs(f["sumaPagos"] - f["valorCancelado"]) > 1
    ]
    print(f"filas donde la suma de meses != VALOR CANCELADO: {len(descuadre)}")
    for f in descuadre:
        print(f"  fila {f['fila']:>3} {f['nombre'][:32]:<32} "
              f"suma={f['sumaPagos']:>14,.0f} cancelado={f['valorCancelado']:>14,.0f}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
