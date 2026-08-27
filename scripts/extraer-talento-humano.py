# -*- coding: utf-8 -*-
"""
Extrae los Excel de talento humano a un JSON que consume
`src/database/scripts/import-talento-humano.ts`.

    python scripts/extraer-talento-humano.py [carpeta-con-los-excel] [salida.json]

Son dos pasos y no uno porque el que sabe leer .xlsx acá es Python (openpyxl) y el
que sabe escribir en la base es el `dataSource` de TypeORM. Este paso **no toca la
base de datos**: solo lee los archivos y escribe un JSON.

Carga inicial: de acá en adelante la información se vive en el sistema y el Excel
deja de ser la fuente.

Tres archivos, tres destinos:

  Base de personal 2026.xlsx           → th_personal      (encabezado en la fila 6)
  INCAPACIDADES.xlsx                   → th_incapacidades (hoja GENERAL)
  01. Ausentismos.xlsx                 → th_ausentismos   (encabezado en la fila 1)
  01. Informe general de préstamos.xlsx→ th_prestamos +
                                         th_prestamo_pagos (hoja «Prestamos»)

**De INCAPACIDADES solo se lee GENERAL**, que es el registro controlado (GTH-31-R).
Las demás hojas del libro quedan por fuera a propósito: 2023/2024/2025 y
«COMP. ELIANA» son borradores ya consolidados en GENERAL —importarlos duplicaría, y
donde difieren es porque GENERAL trae la fecha corregida—, y «jamundi» son 18
registros de esa unión temporal que GENERAL no cubre y que se decidió no traer.

**Del informe de préstamos solo se lee «Prestamos»**, que es la consolidada; las otras
veinte hojas son cortes mensuales de 2023 a 2025 que esa ya recoge.
"""

import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime

import openpyxl

# ── Utilidades de celda ────────────────────────────────────────────────────────

#: Lo que Excel deja cuando una fórmula apunta a una hoja que ya no existe.
BASURA = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


def texto(v, limite=None):
    """Celda → str limpio, o None. El `\\xa0` del Excel cuenta como vacío."""
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if not s or s in BASURA:
        return None
    return s[:limite] if limite else s


def numero(v):
    """Celda → float, o None. Descarta los `#REF!` en vez de propagarlos."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in BASURA:
        return None
    s = re.sub(r"[^\d.,-]", "", s).replace(".", "").replace(",", ".") if s.count(",") else re.sub(r"[^\d.-]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


def entero(v):
    n = numero(v)
    return int(round(n)) if n is not None else None


def iso(v):
    """Celda de fecha → 'YYYY-MM-DD', o None."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def cedula(v):
    """La identificación llega a veces como número: 1112150936.0 no es una cédula."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return texto(v, 30)


# ── El periodo de la incapacidad ───────────────────────────────────────────────

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
    # Como se escribió mal en el archivo. Se aceptan porque el dato es válido
    # aunque la palabra esté torcida.
    "SETIEMBRE": 9, "SEPTEMBRE": 9,
}

_SIN_TILDES = lambda s: "".join(
    c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
)


def _lado(parte):
    """De «08 MAYO 2023» saca (día, mes, año); lo que falte queda en None."""
    mes = next((MESES[p] for p in parte.split() if p in MESES), None)
    anio = next((int(n) for n in re.findall(r"\b(\d{4})\b", parte)), None)
    dia = next((int(n) for n in re.findall(r"\b(\d{1,2})\b", parte)), None)
    return dia, mes, anio


def periodo(txt, respaldo_anio=None):
    """
    «08 MAYO AL 06 JUNIO 2023» → ('2023-05-08', '2023-06-06').

    El archivo escribe el periodo a mano y de seis maneras distintas: con y sin año
    a la izquierda, con el mes solo al final («16 AL 18 ENERO 2025»), con barras
    («07/01/2025 - 13/02/2025») y a veces un solo día («08 DE MAYO 2026»). Se
    devuelve (None, None) cuando de verdad no se deja leer; el texto original
    igual se guarda en `periodo_texto`, así que no se pierde nada.
    """
    if not txt:
        return None, None
    s = _SIN_TILDES(str(txt)).upper().replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    # Formato con barras: 07/01/2025 - 13/02/2025
    barras = re.findall(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", s)
    if barras:
        def de_barras(t):
            d, m, a = int(t[0]), int(t[1]), int(t[2])
            if a < 100:
                a += 2000
            return _fecha(a, m, d)
        ini = de_barras(barras[0])
        fin = de_barras(barras[-1]) if len(barras) > 1 else ini
        return ini, fin

    partes = [p.strip() for p in re.split(r"\bAL\b|\bA\b|-|–", s) if p.strip()]
    if not partes:
        return None, None

    di, mi, ai = _lado(partes[0])
    df, mf, af = _lado(partes[-1]) if len(partes) > 1 else (di, mi, ai)

    # Lo que falta a un lado se toma del otro: «16 AL 18 ENERO 2025» solo trae el
    # mes a la derecha, y «02 AGOSTO 2023 AL 04 AGOSTO 2023» lo trae en los dos.
    mi, mf = mi or mf, mf or mi
    ai, af = ai or af or respaldo_anio, af or ai or respaldo_anio

    ini, fin = _fecha(ai, mi, di), _fecha(af, mf, df)

    # «29 DICIEMBRE AL 05 ENERO 2024»: el año de la derecha no sirve para la
    # izquierda cuando la incapacidad cruza el fin de año.
    if ini and fin and ini > fin:
        ini = _fecha(ai - 1, mi, di)

    return ini, fin


def _fecha(a, m, d):
    """
    Arma la fecha, o None si no cuadra.

    El rango de años no es paranoia: en el archivo hay un «13/02/205» al que le falta
    un dígito, y Postgres acepta feliz el año 205. Una fecha absurda es peor que
    ninguna, porque ninguna se ve y se corrige, y el año 205 se cuela en los informes.
    """
    if not (a and m and d) or not (1990 <= a <= 2100):
        return None
    try:
        return date(a, m, d).isoformat()
    except ValueError:
        return None


# ── Base de personal ───────────────────────────────────────────────────────────

def leer_personal(ruta):
    """
    Hoja única, encabezado en la fila 6, datos de la 7 en adelante.

    Del salario se importa **solo el bloque 2026** (columnas AB..AH). Los de 2024 y
    2025 están llenos de `#REF!` —fórmulas rotas contra hojas que ya no existen—,
    así que traerlos sería traer basura con apariencia de dato.
    """
    ws = openpyxl.load_workbook(ruta, data_only=True)["Hoja1"]
    filas = []
    for r in range(7, ws.max_row + 1):
        c = lambda i: ws.cell(r, i).value  # noqa: E731
        ident, nombre = cedula(c(5)), texto(c(9), 160)
        if not ident or not nombre:
            continue
        filas.append({
            "estado": texto(c(1), 40),
            "tipoContrato": texto(c(2), 60),
            "ubicacion": texto(c(3), 80),
            "empresaProyecto": texto(c(4), 120),
            "identificacion": ident,
            "operacionFge": texto(c(6), 60),
            "centroCosto": texto(c(7), 20),
            "tipoGasto": texto(c(8), 20),
            "nombre": nombre,
            "cargo": texto(c(10), 120),
            "area": texto(c(11), 80),
            "fechaIngreso": iso(c(12)),
            "escalafon": texto(c(13), 80),
            "formacionProfesional": texto(c(14), 80),
            # Bloque 2026: AB=28 … AH=34
            "salario": numero(c(28)),
            "auxilioTransporte": numero(c(29)),
            "auxilioRodamiento": numero(c(30)),
            "totalSalarios": numero(c(31)),
            "cargaPrestacionalPct": numero(c(32)),
            "cargaPrestacional": numero(c(33)),
            "costoTotal": numero(c(34)),
            "anioVigencia": 2026,
        })
    return filas


# ── Incapacidades ──────────────────────────────────────────────────────────────

#: La columna «INCAPACIADAD» del archivo escribe la misma categoría de varias formas
#: —«ACCIDENTE TRABAJO» y «ACCIDENTE DE TRABAJO», «LICENCIA MATENIDAD» con la ene de
#: menos— porque se digitó a mano durante tres años. Se unifican acá, en la puerta: si
#: entran las cinco variantes, el desplegable de la pantalla no puede coincidir con
#: todas y agrupar por tipo parte en dos lo que es un solo tipo.
#:
#: Solo unifica ortografía, nunca junta categorías distintas: un accidente de trabajo y
#: uno de tránsito siguen siendo dos cosas, porque uno lo cubre la ARL y el otro el SOAT.
TIPOS_INCAPACIDAD = {
    "ENFERMEDAD GENERAL": "ENFERMEDAD GENERAL",
    "ACCIDENTE TRABAJO": "ACCIDENTE DE TRABAJO",
    "ACCIDENTE DE TRABAJO": "ACCIDENTE DE TRABAJO",
    "ACCIDENTE TRANSITO": "ACCIDENTE DE TRÁNSITO",
    "ACCIDENTE DE TRANSITO": "ACCIDENTE DE TRÁNSITO",
    "LICENCIA MATENIDAD": "LICENCIA DE MATERNIDAD",
    "LICENCIA MATERNIDAD": "LICENCIA DE MATERNIDAD",
    "LICENCIA DE MATERNIDAD": "LICENCIA DE MATERNIDAD",
    "LICENCIA PATERNIDAD": "LICENCIA DE PATERNIDAD",
    "LICENCIA DE PATERNIDAD": "LICENCIA DE PATERNIDAD",
}


def tipo_incapacidad(v):
    """Normaliza la ortografía del tipo. Lo que no esté en la tabla pasa tal cual."""
    t = texto(v, 80)
    if not t:
        return None
    return TIPOS_INCAPACIDAD.get(_SIN_TILDES(t).upper(), t)


def leer_incapacidades(ruta):
    """
    Hoja GENERAL, encabezado en la fila 6, datos de la 7 en adelante.

    Ojo con dos columnas del archivo que no dicen lo que parecen:

    - La **P**, rotulada «ESTADO», no trae el estado sino el radicado —o la razón
      por la que no se pudo radicar—. Va a `numero_radicacion` cuando es un
      radicado y a `observaciones` cuando es una explicación.
    - La **U**, rotulada «Esatdo», sí es el estado normalizado, y es la que manda.
    """
    ws = openpyxl.load_workbook(ruta, data_only=True)["GENERAL"]
    filas = []
    for r in range(7, ws.max_row + 1):
        c = lambda i: ws.cell(r, i).value  # noqa: E731
        ident, nombre = cedula(c(2)), texto(c(3), 160)
        if not ident or not nombre:
            continue

        marca = iso(c(20))  # col T: fecha suelta, a veces solo el mes
        texto_periodo = texto(c(11), 120)
        ini, fin = periodo(texto_periodo, int(marca[:4]) if marca else None)

        radicado = texto(c(16))
        es_radicado = bool(radicado) and (
            re.search(r"RADICAD", radicado, re.I) or re.fullmatch(r"[\w.-]{4,}", radicado)
        )

        filas.append({
            "identificacion": ident,
            "nombre": nombre,
            "proyecto": texto(c(4), 120),
            "salario": numero(c(5)),
            "tipo": tipo_incapacidad(c(6)),
            "tipoAfectacion": texto(c(7), 80),
            "totalDias": entero(c(8)),
            "diasEmpresa": entero(c(9)),
            "diasEntidad": entero(c(10)),
            "periodoTexto": texto_periodo,
            "fechaInicio": ini or marca,
            "fechaFin": fin,
            "valorAsumidoEmpresa": numero(c(12)),
            "valorRecobro": numero(c(13)),
            "entidad": texto(c(14), 120),
            "numeroIncapacidad": texto(c(15), 60),
            "numeroRadicacion": radicado[:80] if es_radicado else None,
            "valorProyectadoRecuperar": numero(c(17)),
            "valorRecuperado": numero(c(18)),
            "fechaPago": texto(c(19), 80),
            "estado": texto(c(21), 60),
            "observaciones": None if es_radicado else radicado,
        })
    return filas


# ── Ausentismos ────────────────────────────────────────────────────────────────

def hora(v):
    """Celda de hora → 'HH:MM', o None."""
    if v is None:
        return None
    if hasattr(v, "hour"):
        return "%02d:%02d" % (v.hour, v.minute)
    s = texto(v)
    m = re.match(r"^(\d{1,2}):(\d{2})", s or "")
    return "%02d:%02d" % (int(m.group(1)), int(m.group(2))) if m else None


def leer_ausentismos(ruta):
    """Encabezado en la fila 1, datos de la 2 en adelante."""
    ws = openpyxl.load_workbook(ruta, data_only=True)["AUSENTISMO "]
    filas = []
    for r in range(2, ws.max_row + 1):
        c = lambda i: ws.cell(r, i).value  # noqa: E731
        ident, nombre = cedula(c(2)), texto(c(3), 160)
        if not ident or not nombre:
            continue
        filas.append({
            "identificacion": ident,
            "nombre": nombre,
            "cargo": texto(c(4), 120),
            "area": texto(c(5), 80),
            "tipoContrato": texto(c(6), 60),
            "fechaInicio": iso(c(7)),
            "fechaFin": iso(c(8)),
            "horaSalida": hora(c(9)),
            "horaEntrada": hora(c(10)),
            "horasAusencia": numero(c(11)),
            "diasPermiso": entero(c(12)),
            "motivo": texto(c(13), 120),
            "soporte": texto(c(14), 40),
            "observaciones": texto(c(15)),
        })
    return filas


# ── Préstamos ──────────────────────────────────────────────────────────────────

def indice_por_nombre(personal):
    """
    Nombre normalizado → cédula, para poder resolver a quién es cada préstamo.

    Se indexa por el **conjunto de palabras** y no por la cadena: la base de personal
    escribe «BAEZA MARÍN YAKI MICHELL» (apellidos primero, en mayúscula y con tildes) y la
    hoja de préstamos «Yamile Rodriguez» (nombre primero, sin tildes y a veces sin el
    segundo apellido). Como conjunto, «MINA MINA JOSE ARNU» y «Jose Arnu Mina» coinciden.

    Los nombres que aparecen dos veces en la base —alguien que salió y volvió a entrar—
    se descartan del índice: si no se sabe a cuál de los dos registros va, es mejor dejar
    la cédula vacía que colgarle el préstamo al equivocado.
    """
    por_clave = {}
    for p in personal:
        clave = frozenset(_SIN_TILDES(p["nombre"]).upper().split())
        if not clave:
            continue
        por_clave.setdefault(clave, set()).add(p["identificacion"])
    return {k: next(iter(v)) for k, v in por_clave.items() if len(v) == 1}


def buscar_cedula(nombre, indice):
    """Coincidencia exacta, o por subconjunto cuando es la única candidata."""
    palabras = frozenset(_SIN_TILDES(nombre).upper().split())
    if not palabras:
        return None
    if palabras in indice:
        return indice[palabras]
    # «Jose Arnu Mina» ⊂ «MINA MINA JOSE ARNU»: se acepta solo si nadie más la contiene.
    candidatas = {ced for clave, ced in indice.items() if palabras <= clave}
    return candidatas.pop() if len(candidatas) == 1 else None


def cuotas(v):
    """
    El «# CUOTAS» puede llegar como fecha.

    Una fila quedó con formato de fecha y Excel muestra el 10 como 1900-01-10; el serial
    de Excel arranca el 1900-01-01, así que el día del mes es el número que se digitó.
    """
    if isinstance(v, (datetime, date)):
        return v.day if v.year == 1900 else None
    return entero(v)


#: La retícula de pagos: columnas J..BI, un mes cada una.
#: J..M son los cuatro últimos meses de 2022 y de ahí en adelante corren doce por año.
PRIMERA_COL_PAGO, ULTIMA_COL_PAGO = 10, 61


def _mes_de_columna(c):
    """Columna de la retícula → (año, mes)."""
    if c <= 13:                       # J..M = septiembre a diciembre de 2022
        return 2022, 9 + (c - 10)
    desplazamiento = c - 14           # N = enero de 2023
    return 2023 + desplazamiento // 12, desplazamiento % 12 + 1


def leer_prestamos(ruta, personal):
    """
    Hoja «Prestamos»: encabezado en la fila 2, meses en la 3, datos de la 4 en adelante.

    Se corta en cuanto la columna A dice «TOTALES»: debajo vienen los totales y unos
    bloques de cuadre que no son préstamos.
    """
    ws = openpyxl.load_workbook(ruta, data_only=True)["Prestamos"]
    indice = indice_por_nombre(personal)
    h64 = _SIN_TILDES(texto(ws.cell(4, 64).value) or "").upper()
    h65 = _SIN_TILDES(texto(ws.cell(4, 65).value) or "").upper()
    tiene_descuento_nomina = "NOMBRE" in h64 and "CUOTA" in h65

    prestamos, sin_cedula = [], 0
    for r in range(4, ws.max_row + 1):
        c = lambda i: ws.cell(r, i).value  # noqa: E731

        if texto(c(1)) and "TOTAL" in _SIN_TILDES(texto(c(1))).upper():
            break
        nombre = texto(c(2), 160)
        if not nombre:
            continue

        cedula_hallada = buscar_cedula(nombre, indice)
        if not cedula_hallada:
            sin_cedula += 1

        # La retícula se endereza a filas y se descartan los ceros: en la hoja son
        # «este mes no se descontó», no un pago de cero pesos.
        pagos = []
        for col in range(PRIMERA_COL_PAGO, ULTIMA_COL_PAGO + 1):
            valor = numero(c(col))
            if valor:
                anio, mes = _mes_de_columna(col)
                pagos.append({"anio": anio, "mes": mes, "valor": valor})

        prestamos.append({
            "numero": entero(c(1)),
            "nombre": nombre,
            "identificacion": cedula_hallada,
            "proyecto": texto(c(3), 20),
            "pagare": (texto(c(4), 10) or "").upper() or None,
            "mesInicio": iso(c(5)),
            "numeroCuotas": cuotas(c(6)),
            "fechaVencimiento": iso(c(7)),
            "valorPrestamo": numero(c(8)),
            "valorCuota": numero(c(9)),
            "valorCancelado": numero(c(62)),   # BJ
            "saldo": numero(c(63)),            # BK
            "nombreNomina": texto(c(64), 160) if tiene_descuento_nomina else None,  # BL
            "cuotaDescontar": numero(c(65)) if tiene_descuento_nomina else None,    # BM
            "observaciones": None if tiene_descuento_nomina else texto(c(64)),      # BL
            "pagos": pagos,
        })

    if sin_cedula:
        print("      (%d de %d préstamos quedaron sin cédula: el nombre no cuadró"
              " con la base de personal)" % (sin_cedula, len(prestamos)))
    return prestamos


# ── Entrada ────────────────────────────────────────────────────────────────────

def main():
    carpeta = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.getcwd(), "..")
    salida = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.getcwd(), "talento-humano.json")

    datos = {}

    def leer(clave, archivo, funcion, *extra):
        ruta = os.path.join(carpeta, archivo)
        if not os.path.exists(ruta):
            print("  falta %s — se omite" % archivo)
            datos[clave] = []
            return []
        datos[clave] = funcion(ruta, *extra)
        print("  %-14s %4d registros  (%s)" % (clave, len(datos[clave]), archivo))
        return datos[clave]

    # El personal va de primero porque los préstamos lo usan para resolver la cédula:
    # su hoja solo trae el nombre.
    personal = leer("personal", "Base de personal 2026.xlsx", leer_personal)
    leer("incapacidades", "INCAPACIDADES.xlsx", leer_incapacidades)
    leer("ausentismos", "01. Ausentismos.xlsx", leer_ausentismos)
    prestamos = leer("prestamos", "01. Informe general de préstamos.xlsx",
                     leer_prestamos, personal)

    if prestamos:
        print("  %-14s %4d cuotas descontadas" % ("(pagos)",
              sum(len(p["pagos"]) for p in prestamos)))

    with open(salida, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=1)
    print("\nJSON escrito en %s" % salida)


if __name__ == "__main__":
    main()
